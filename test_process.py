import sys
sys.path.insert(0, '.')
from app import *
import openpyxl

perf_wb = openpyxl.load_workbook(r'C:\Users\ReubenG\Downloads\xlsv\Performance - OVERALL.xlsx', data_only=True)
sa_wb = openpyxl.load_workbook(r'C:\Users\ReubenG\Downloads\xlsv\Scheme-A.xlsx', data_only=True)
sb_wb = openpyxl.load_workbook(r'C:\Users\ReubenG\Downloads\xlsv\Scheme-B.xlsx', data_only=True)

print('Parsing Performance...')
employees, date_labels = parse_performance(perf_wb)
print(f'  Employees: {len(employees)}')
print(f'  Date labels: {date_labels}')
emp0 = employees[0]
print(f'  Sample: code={emp0["code"]}, id={emp0["employee_id"]}, name={emp0["name"]}')
print(f'  Shifts: {emp0["shift"][:10]}')
print(f'  Leave: {emp0["leave"][:10]}')

full_dates = resolve_dates(date_labels)
print(f'  Full dates: {full_dates[:5]} ... {full_dates[-3:]}')

print('\nParsing Scheme-A...')
sa_entries, sa_dates = parse_scheme_a(sa_wb)
print(f'  Entries: {len(sa_entries)}, Dates: {sa_dates[:5]}')

print('\nParsing Scheme-B...')
sb_entries = parse_scheme_b(sb_wb)
print(f'  Entries: {len(sb_entries)}, Sample: {sb_entries[0]}')

print('\nBuilding output...')
output_wb = build_output(perf_wb, sa_wb, sb_wb)
print(f'  Sheets: {output_wb.sheetnames}')
for sn in output_wb.sheetnames:
    ws = output_wb[sn]
    print(f'  {sn}: {ws.max_row} rows x {ws.max_column} cols')

output_wb.save('test_output.xlsx')
print('\nSUCCESS! Saved test_output.xlsx')

perf_wb.close()
sa_wb.close()
sb_wb.close()
