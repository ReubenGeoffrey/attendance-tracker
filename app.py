import os
import io
import json
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, g, jsonify
)
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

app = Flask(__name__)
app.secret_key = "shift-processor-secret-key-2026"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB max upload

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ── Hardcoded admin credentials ──────────────────────────────────
ADMIN_USER = "admin"
ADMIN_PASS = generate_password_hash("admin123")

# Shift code mapping
SHIFT_MAP = {
    "AA": "A",
    "BB": "B",
    "CC": "C",
    "GG": "G",   # General
    "RR": "R",   # Rest
    "WO": "WO",  # Week Off
}


# ── Auth ─────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "logged_in" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/")
def index():
    if "logged_in" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username == ADMIN_USER and check_password_hash(ADMIN_PASS, password):
            session["logged_in"] = True
            session["username"] = username
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Dashboard ────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")


# ── Upload & Process ─────────────────────────────────────────────
@app.route("/process", methods=["POST"])
@login_required
def process():
    perf_file = request.files.get("performance")
    scheme_a_file = request.files.get("scheme_a")
    scheme_b_file = request.files.get("scheme_b")

    if not perf_file or not scheme_a_file or not scheme_b_file:
        flash("Please upload all 3 files.", "error")
        return redirect(url_for("dashboard"))

    try:
        perf_wb = openpyxl.load_workbook(perf_file, data_only=True)
        sa_wb = openpyxl.load_workbook(scheme_a_file, data_only=True)
        sb_wb = openpyxl.load_workbook(scheme_b_file, data_only=True)

        output_wb = build_output(perf_wb, sa_wb, sb_wb)

        perf_wb.close()
        sa_wb.close()
        sb_wb.close()

        buf = io.BytesIO()
        output_wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"Processed_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Processing error: {str(e)}", "error")
        return redirect(url_for("dashboard"))


# ── Preview endpoint (AJAX) ──────────────────────────────────────
@app.route("/preview", methods=["POST"])
@login_required
def preview():
    perf_file = request.files.get("performance")
    scheme_a_file = request.files.get("scheme_a")
    scheme_b_file = request.files.get("scheme_b")

    if not perf_file or not scheme_a_file or not scheme_b_file:
        return jsonify({"error": "Please upload all 3 files."}), 400

    try:
        perf_wb = openpyxl.load_workbook(perf_file, data_only=True)
        sa_wb = openpyxl.load_workbook(scheme_a_file, data_only=True)
        sb_wb = openpyxl.load_workbook(scheme_b_file, data_only=True)

        result = build_preview(perf_wb, sa_wb, sb_wb)

        perf_wb.close()
        sa_wb.close()
        sb_wb.close()

        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  CORE PROCESSING LOGIC
# ══════════════════════════════════════════════════════════════════

def parse_performance(perf_wb):
    """Parse Performance sheet into structured employee data."""
    ws = perf_wb["Performance"]
    rows = list(ws.iter_rows(values_only=True))

    # Header row (row 1) has date columns at index 4..34
    header = rows[0]
    date_labels = [str(header[i]).zfill(2) if header[i] else "" for i in range(4, 35)]

    # Sheet1 has CODE list — build code-to-id mapping
    code_list = []
    if "Sheet1" in perf_wb.sheetnames:
        ws1 = perf_wb["Sheet1"]
        for r in ws1.iter_rows(min_row=2, values_only=True):
            if r[0] is not None:
                code_list.append(r[0])

    # Parse employees (11 rows per employee block)
    FIELDS = ["INTIME", "OUTTIME", "LATE", "EARLY", "OT",
              "SHIFT", "LEAVE", "WORKHRS", "COREHRS", "LOGINHRS", "SHORTWORKHRS"]
    employees = []
    i = 1  # skip header
    while i < len(rows):
        row = rows[i]
        if row[1] is not None:  # Code column has value → start of employee block
            emp = {
                "sr_no": row[0],
                "code": row[1],
                "name": row[2],
                "employee_id": try_numeric(row[1]),
            }
            # Parse 11 field rows
            for fi, field in enumerate(FIELDS):
                field_row = rows[i + fi] if (i + fi) < len(rows) else [None] * 66
                vals = []
                for ci in range(4, 35):
                    v = field_row[ci] if ci < len(field_row) else None
                    vals.append(v if v is not None and v != "" else None)
                emp[field.lower()] = vals

                # Also grab summary columns (index 35+) from first row only
                if fi == 0:
                    summary_keys = list(header[35:]) if len(header) > 35 else []
                    for si, sk in enumerate(summary_keys):
                        idx = 35 + si
                        emp[f"summary_{sk}"] = field_row[idx] if idx < len(field_row) else None

            employees.append(emp)
            i += len(FIELDS)
        else:
            i += 1

    return employees, date_labels


def parse_scheme_a(sa_wb):
    """Parse Scheme-A Addition sheet."""
    ws = sa_wb["Addition"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 1 = day names, Row 2 = headers with dates
    header_row = rows[1]  # EMP.NO, NAME, SECTION, date1, date2, ...
    dates = []
    for i in range(3, len(header_row)):
        v = header_row[i]
        if isinstance(v, datetime):
            dates.append(v.strftime("%Y-%m-%d"))
        elif v:
            dates.append(str(v))
        else:
            dates.append(None)

    entries = []
    for r in rows[2:]:
        if r[0] is None:
            continue
        emp_no = r[0]
        name = r[1]
        section = r[2]
        efficiencies = []
        for i in range(3, len(r)):
            v = r[i]
            if v is not None and v != "":
                try:
                    efficiencies.append(round(float(v), 2))
                except (ValueError, TypeError):
                    efficiencies.append(None)
            else:
                efficiencies.append(None)
        entries.append({
            "emp_no": emp_no,
            "employee_id": try_numeric(emp_no),
            "name": name,
            "section": section,
            "efficiencies": efficiencies,
        })

    return entries, dates


def parse_scheme_b(sb_wb):
    """Parse Scheme-B sheet."""
    ws = sb_wb[sb_wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    entries = []
    for r in rows:
        # Find rows with Date in col D (index 3) that are datetime
        if r[3] and isinstance(r[3], datetime):
            date_str = r[3].strftime("%Y-%m-%d")
            a_val = safe_float(r[4])
            b_val = safe_float(r[5])
            c_val = safe_float(r[6])
            avg_val = safe_float(r[7])
            entries.append({
                "date": date_str,
                "A": a_val,
                "B": b_val,
                "C": c_val,
                "AVG": avg_val,
            })

    return entries


def try_numeric(val):
    """Convert value to numeric if possible."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return s


def safe_float(val):
    """Safely convert to float, return 0 if not possible."""
    if val is None or val == "":
        return 0.0
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return 0.0


def resolve_dates(date_labels):
    """Convert day-of-month labels (19..31, 01..18) to full dates.
    Assumes payroll month structure: starts mid-month."""
    # We need to figure out the year/month from context
    # Labels: 19,20,...,31,01,02,...,18
    # The first part (19-31) is one month, the second part (01-18) is the next month
    now = datetime.now()
    # Try to infer: if current month day <= 18, the period ended this month
    # Otherwise it ends next month
    year = now.year

    # Build date list
    dates = []
    first_month = None
    second_month = None

    # Find the split point: where labels go from high to low
    split_idx = 0
    for i in range(1, len(date_labels)):
        try:
            prev = int(date_labels[i - 1])
            curr = int(date_labels[i])
            if curr < prev:
                split_idx = i
                break
        except ValueError:
            continue

    # Use current context to determine months
    # Default: assume current payroll period
    first_day = int(date_labels[0]) if date_labels[0] else 19
    if first_day >= 19:
        # First chunk is previous month, second chunk is current month
        month2 = now.month
        month1 = month2 - 1 if month2 > 1 else 12
        year1 = year if month2 > 1 else year - 1
        year2 = year
    else:
        month1 = now.month
        month2 = month1 + 1 if month1 < 12 else 1
        year1 = year
        year2 = year if month1 < 12 else year + 1

    for i, lbl in enumerate(date_labels):
        if not lbl:
            dates.append("")
            continue
        try:
            day = int(lbl)
            if i < split_idx:
                dates.append(f"{year1}-{month1:02d}-{day:02d}")
            else:
                dates.append(f"{year2}-{month2:02d}-{day:02d}")
        except ValueError:
            dates.append("")

    return dates


# ══════════════════════════════════════════════════════════════════
#  BUILD OUTPUT WORKBOOK
# ══════════════════════════════════════════════════════════════════

# Shared styles
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Segoe UI", size=9)
CELL_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill(start_color="F0F0F5", end_color="F0F0F5", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, row_num, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row_num, col_num, alt=False):
    cell = ws.cell(row=row_num, column=col_num)
    cell.font = CELL_FONT
    cell.alignment = CELL_ALIGN
    cell.border = THIN_BORDER
    if alt:
        cell.fill = ALT_FILL


def build_output(perf_wb, sa_wb, sb_wb):
    """Build the complete output workbook with 8 sheets."""
    employees, date_labels = parse_performance(perf_wb)
    sa_entries, sa_dates = parse_scheme_a(sa_wb)
    sb_entries = parse_scheme_b(sb_wb)
    full_dates = resolve_dates(date_labels)

    # Build Scheme-B lookup: date -> {A, B, C, AVG}
    sb_lookup = {}
    for entry in sb_entries:
        sb_lookup[entry["date"]] = entry

    # Build Scheme-A lookup: emp_no -> entry
    sa_lookup = {}
    for entry in sa_entries:
        sa_lookup[str(entry["emp_no"])] = entry

    wb = openpyxl.Workbook()

    # ── Sheet 1: SHIFT ────────────────────────────────────────
    ws_shift = wb.active
    ws_shift.title = "SHIFT"
    headers = ["EMPLOYEE_ID", "CODE", "NAME"] + [d if d else f"Day{i}" for i, d in enumerate(full_dates)]
    for c, h in enumerate(headers, 1):
        ws_shift.cell(row=1, column=c, value=h)
    style_header_row(ws_shift, 1, len(headers))

    for ri, emp in enumerate(employees, 2):
        ws_shift.cell(row=ri, column=1, value=emp["employee_id"])
        ws_shift.cell(row=ri, column=2, value=emp["code"])
        ws_shift.cell(row=ri, column=3, value=emp["name"])
        for di, shift_val in enumerate(emp.get("shift", [])):
            ws_shift.cell(row=ri, column=4 + di, value=shift_val)
        alt = ri % 2 == 0
        for c in range(1, len(headers) + 1):
            style_data_cell(ws_shift, ri, c, alt)

    ws_shift.column_dimensions["A"].width = 14
    ws_shift.column_dimensions["B"].width = 10
    ws_shift.column_dimensions["C"].width = 22
    for ci in range(4, len(headers) + 1):
        ws_shift.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12

    # ── Sheet 2: A&B ──────────────────────────────────────────
    ws_ab = wb.create_sheet("A&B")
    ab_headers = ["EMPLOYEE_ID", "CODE", "NAME"] + [d if d else f"Day{i}" for i, d in enumerate(full_dates)]
    for c, h in enumerate(ab_headers, 1):
        ws_ab.cell(row=1, column=c, value=h)
    style_header_row(ws_ab, 1, len(ab_headers))

    for ri, emp in enumerate(employees, 2):
        ws_ab.cell(row=ri, column=1, value=emp["employee_id"])
        ws_ab.cell(row=ri, column=2, value=emp["code"])
        ws_ab.cell(row=ri, column=3, value=emp["name"])
        leave_vals = emp.get("leave", [])
        shift_vals = emp.get("shift", [])
        for di in range(len(full_dates)):
            leave = leave_vals[di] if di < len(leave_vals) else None
            shift = shift_vals[di] if di < len(shift_vals) else None
            # Show leave status + shift info
            if leave and leave not in ("P", None):
                ws_ab.cell(row=ri, column=4 + di, value=leave)
            elif shift:
                ws_ab.cell(row=ri, column=4 + di, value=shift)
        alt = ri % 2 == 0
        for c in range(1, len(ab_headers) + 1):
            style_data_cell(ws_ab, ri, c, alt)

    ws_ab.column_dimensions["A"].width = 14
    ws_ab.column_dimensions["B"].width = 10
    ws_ab.column_dimensions["C"].width = 22

    # ── Sheet 3: CF (Curing Factor — Scheme B Efficiency) ─────
    ws_cf = wb.create_sheet("CF")
    cf_headers = ["DATE", "SHIFT_A", "SHIFT_B", "SHIFT_C", "AVG"]
    for c, h in enumerate(cf_headers, 1):
        ws_cf.cell(row=1, column=c, value=h)
    style_header_row(ws_cf, 1, len(cf_headers))

    # Use all dates from full_dates, fill from scheme-B lookup
    all_sb_dates = sorted(set(full_dates) | set(sb_lookup.keys()))
    all_sb_dates = [d for d in all_sb_dates if d]

    for ri, d in enumerate(all_sb_dates, 2):
        ws_cf.cell(row=ri, column=1, value=d)
        if d in sb_lookup:
            ws_cf.cell(row=ri, column=2, value=sb_lookup[d]["A"])
            ws_cf.cell(row=ri, column=3, value=sb_lookup[d]["B"])
            ws_cf.cell(row=ri, column=4, value=sb_lookup[d]["C"])
            ws_cf.cell(row=ri, column=5, value=sb_lookup[d]["AVG"])
        else:
            # No production — mark as 0
            ws_cf.cell(row=ri, column=2, value=0)
            ws_cf.cell(row=ri, column=3, value=0)
            ws_cf.cell(row=ri, column=4, value=0)
            ws_cf.cell(row=ri, column=5, value=0)
        alt = ri % 2 == 0
        for c in range(1, 6):
            style_data_cell(ws_cf, ri, c, alt)

    ws_cf.column_dimensions["A"].width = 14
    for col in ["B", "C", "D", "E"]:
        ws_cf.column_dimensions[col].width = 14

    # ── Sheet 4: Sheet A (Scheme A — Employee ID & Date) ──────
    ws_sa = wb.create_sheet("Sheet A")
    sa_headers_out = ["EMPLOYEE_ID", "EMP_NO", "NAME", "SECTION"] + sa_dates
    for c, h in enumerate(sa_headers_out, 1):
        ws_sa.cell(row=1, column=c, value=h)
    style_header_row(ws_sa, 1, len(sa_headers_out))

    for ri, entry in enumerate(sa_entries, 2):
        ws_sa.cell(row=ri, column=1, value=entry["employee_id"])
        ws_sa.cell(row=ri, column=2, value=entry["emp_no"])
        ws_sa.cell(row=ri, column=3, value=entry["name"])
        ws_sa.cell(row=ri, column=4, value=entry["section"])
        for di, eff in enumerate(entry["efficiencies"]):
            ws_sa.cell(row=ri, column=5 + di, value=eff)
        alt = ri % 2 == 0
        for c in range(1, len(sa_headers_out) + 1):
            style_data_cell(ws_sa, ri, c, alt)

    ws_sa.column_dimensions["A"].width = 14
    ws_sa.column_dimensions["B"].width = 12
    ws_sa.column_dimensions["C"].width = 22
    ws_sa.column_dimensions["D"].width = 20

    # ── Sheet 5: Sheet B (Scheme B by reporting shift) ────────
    ws_sb = wb.create_sheet("Sheet B")
    sb_headers = ["EMPLOYEE_ID", "CODE", "NAME"] + [d if d else f"Day{i}" for i, d in enumerate(full_dates)]
    for c, h in enumerate(sb_headers, 1):
        ws_sb.cell(row=1, column=c, value=h)
    style_header_row(ws_sb, 1, len(sb_headers))

    for ri, emp in enumerate(employees, 2):
        ws_sb.cell(row=ri, column=1, value=emp["employee_id"])
        ws_sb.cell(row=ri, column=2, value=emp["code"])
        ws_sb.cell(row=ri, column=3, value=emp["name"])
        shift_vals = emp.get("shift", [])
        for di in range(len(full_dates)):
            date_str = full_dates[di] if di < len(full_dates) else ""
            shift_code = shift_vals[di] if di < len(shift_vals) else None
            mapped_shift = SHIFT_MAP.get(shift_code, "") if shift_code else ""

            if date_str and date_str in sb_lookup and mapped_shift in ("A", "B", "C"):
                eff = sb_lookup[date_str].get(mapped_shift, 0)
                ws_sb.cell(row=ri, column=4 + di, value=round(eff, 2) if eff else 0)
            elif mapped_shift in ("WO", "R", "G"):
                ws_sb.cell(row=ri, column=4 + di, value=mapped_shift)
            else:
                ws_sb.cell(row=ri, column=4 + di, value=0)
        alt = ri % 2 == 0
        for c in range(1, len(sb_headers) + 1):
            style_data_cell(ws_sb, ri, c, alt)

    ws_sb.column_dimensions["A"].width = 14
    ws_sb.column_dimensions["B"].width = 10
    ws_sb.column_dimensions["C"].width = 22

    # ── Sheet 6: OP (Output — Operational/Present Data) ───────
    ws_op = wb.create_sheet("OP")
    op_headers = ["EMPLOYEE_ID", "CODE", "NAME", "DATE", "SHIFT", "INTIME", "OUTTIME", "WORKHRS", "OT", "STATUS"]
    for c, h in enumerate(op_headers, 1):
        ws_op.cell(row=1, column=c, value=h)
    style_header_row(ws_op, 1, len(op_headers))

    op_row = 2
    for emp in employees:
        shift_vals = emp.get("shift", [])
        leave_vals = emp.get("leave", [])
        intime_vals = emp.get("intime", [])
        outtime_vals = emp.get("outtime", [])
        workhrs_vals = emp.get("workhrs", [])
        ot_vals = emp.get("ot", [])
        for di in range(len(full_dates)):
            date_str = full_dates[di] if di < len(full_dates) else ""
            leave = leave_vals[di] if di < len(leave_vals) else None
            if leave == "P" or (leave is None and intime_vals[di] if di < len(intime_vals) else False):
                ws_op.cell(row=op_row, column=1, value=emp["employee_id"])
                ws_op.cell(row=op_row, column=2, value=emp["code"])
                ws_op.cell(row=op_row, column=3, value=emp["name"])
                ws_op.cell(row=op_row, column=4, value=date_str)
                ws_op.cell(row=op_row, column=5, value=shift_vals[di] if di < len(shift_vals) else "")
                ws_op.cell(row=op_row, column=6, value=intime_vals[di] if di < len(intime_vals) else "")
                ws_op.cell(row=op_row, column=7, value=outtime_vals[di] if di < len(outtime_vals) else "")
                ws_op.cell(row=op_row, column=8, value=workhrs_vals[di] if di < len(workhrs_vals) else "")
                ws_op.cell(row=op_row, column=9, value=ot_vals[di] if di < len(ot_vals) else "")
                ws_op.cell(row=op_row, column=10, value="Present")
                alt = op_row % 2 == 0
                for c in range(1, len(op_headers) + 1):
                    style_data_cell(ws_op, op_row, c, alt)
                op_row += 1

    for col_letter, w in [("A", 14), ("B", 10), ("C", 22), ("D", 12), ("E", 8),
                           ("F", 10), ("G", 10), ("H", 10), ("I", 8), ("J", 10)]:
        ws_op.column_dimensions[col_letter].width = w

    # ── Sheet 7: CL (Leave Output) ───────────────────────────
    ws_cl = wb.create_sheet("CL")
    cl_headers = ["EMPLOYEE_ID", "CODE", "NAME", "DATE", "SHIFT", "LEAVE_TYPE"]
    for c, h in enumerate(cl_headers, 1):
        ws_cl.cell(row=1, column=c, value=h)
    style_header_row(ws_cl, 1, len(cl_headers))

    cl_row = 2
    for emp in employees:
        shift_vals = emp.get("shift", [])
        leave_vals = emp.get("leave", [])
        for di in range(len(full_dates)):
            date_str = full_dates[di] if di < len(full_dates) else ""
            leave = leave_vals[di] if di < len(leave_vals) else None
            if leave and leave not in ("P", None) and leave != "WO":
                ws_cl.cell(row=cl_row, column=1, value=emp["employee_id"])
                ws_cl.cell(row=cl_row, column=2, value=emp["code"])
                ws_cl.cell(row=cl_row, column=3, value=emp["name"])
                ws_cl.cell(row=cl_row, column=4, value=date_str)
                ws_cl.cell(row=cl_row, column=5, value=shift_vals[di] if di < len(shift_vals) else "")
                ws_cl.cell(row=cl_row, column=6, value=leave)
                alt = cl_row % 2 == 0
                for c in range(1, len(cl_headers) + 1):
                    style_data_cell(ws_cl, cl_row, c, alt)
                cl_row += 1

    for col_letter, w in [("A", 14), ("B", 10), ("C", 22), ("D", 12), ("E", 8), ("F", 12)]:
        ws_cl.column_dimensions[col_letter].width = w

    # ── Sheet 8: DAT (Full Data Output) ──────────────────────
    ws_dat = wb.create_sheet("DAT")
    dat_headers = ["EMPLOYEE_ID", "CODE", "NAME", "DATE", "SHIFT", "LEAVE",
                   "INTIME", "OUTTIME", "WORKHRS", "LATE", "EARLY", "OT",
                   "COREHRS", "LOGINHRS", "SHORTWORKHRS"]
    for c, h in enumerate(dat_headers, 1):
        ws_dat.cell(row=1, column=c, value=h)
    style_header_row(ws_dat, 1, len(dat_headers))

    dat_row = 2
    for emp in employees:
        for di in range(len(full_dates)):
            date_str = full_dates[di] if di < len(full_dates) else ""
            if not date_str:
                continue
            ws_dat.cell(row=dat_row, column=1, value=emp["employee_id"])
            ws_dat.cell(row=dat_row, column=2, value=emp["code"])
            ws_dat.cell(row=dat_row, column=3, value=emp["name"])
            ws_dat.cell(row=dat_row, column=4, value=date_str)
            fields_map = {
                5: "shift", 6: "leave", 7: "intime", 8: "outtime",
                9: "workhrs", 10: "late", 11: "early", 12: "ot",
                13: "corehrs", 14: "loginhrs", 15: "shortworkhrs"
            }
            for col, field in fields_map.items():
                vals = emp.get(field, [])
                ws_dat.cell(row=dat_row, column=col, value=vals[di] if di < len(vals) else "")
            alt = dat_row % 2 == 0
            for c in range(1, len(dat_headers) + 1):
                style_data_cell(ws_dat, dat_row, c, alt)
            dat_row += 1

    for col_letter, w in [("A", 14), ("B", 10), ("C", 22), ("D", 12), ("E", 8),
                           ("F", 8), ("G", 10), ("H", 10), ("I", 10), ("J", 10),
                           ("K", 10), ("L", 8), ("M", 10), ("N", 10), ("O", 14)]:
        ws_dat.column_dimensions[col_letter].width = w

    return wb


def build_preview(perf_wb, sa_wb, sb_wb):
    """Build preview data (first 20 rows per sheet) for AJAX response."""
    employees, date_labels = parse_performance(perf_wb)
    sa_entries, sa_dates = parse_scheme_a(sa_wb)
    sb_entries = parse_scheme_b(sb_wb)
    full_dates = resolve_dates(date_labels)
    sb_lookup = {e["date"]: e for e in sb_entries}

    preview = {}
    LIMIT = 20

    # SHIFT preview
    shift_headers = ["EMPLOYEE_ID", "CODE", "NAME"] + full_dates[:10] + ["..."]
    shift_rows = []
    for emp in employees[:LIMIT]:
        row = [emp["employee_id"], emp["code"], emp["name"]]
        for s in (emp.get("shift", [])[:10]):
            row.append(s if s else "")
        row.append("...")
        shift_rows.append(row)
    preview["SHIFT"] = {"headers": shift_headers, "rows": shift_rows, "total": len(employees)}

    # CF preview
    cf_headers = ["DATE", "SHIFT_A", "SHIFT_B", "SHIFT_C", "AVG"]
    cf_rows = []
    for e in sb_entries[:LIMIT]:
        cf_rows.append([e["date"], e["A"], e["B"], e["C"], e["AVG"]])
    preview["CF"] = {"headers": cf_headers, "rows": cf_rows, "total": len(sb_entries)}

    # Sheet A preview
    sa_h = ["EMPLOYEE_ID", "NAME", "SECTION"] + sa_dates[:8] + ["..."]
    sa_rows = []
    for e in sa_entries[:LIMIT]:
        row = [e["employee_id"], e["name"], e["section"]]
        for eff in e["efficiencies"][:8]:
            row.append(eff if eff is not None else "")
        row.append("...")
        sa_rows.append(row)
    preview["Sheet A"] = {"headers": sa_h, "rows": sa_rows, "total": len(sa_entries)}

    # Sheet B preview
    sb_h = ["EMPLOYEE_ID", "NAME"] + full_dates[:8] + ["..."]
    sb_rows = []
    for emp in employees[:LIMIT]:
        row = [emp["employee_id"], emp["name"]]
        shift_vals = emp.get("shift", [])
        for di in range(min(8, len(full_dates))):
            date_str = full_dates[di]
            shift_code = shift_vals[di] if di < len(shift_vals) else None
            mapped = SHIFT_MAP.get(shift_code, "") if shift_code else ""
            if date_str in sb_lookup and mapped in ("A", "B", "C"):
                row.append(round(sb_lookup[date_str][mapped], 2))
            else:
                row.append(mapped or 0)
        row.append("...")
        sb_rows.append(row)
    preview["Sheet B"] = {"headers": sb_h, "rows": sb_rows, "total": len(employees)}

    # OP preview
    op_h = ["EMPLOYEE_ID", "NAME", "DATE", "SHIFT", "INTIME", "OUTTIME", "WORKHRS"]
    op_rows = []
    count_op = 0
    for emp in employees:
        leave_vals = emp.get("leave", [])
        for di in range(len(full_dates)):
            leave = leave_vals[di] if di < len(leave_vals) else None
            if leave == "P":
                count_op += 1
                if len(op_rows) < LIMIT:
                    op_rows.append([
                        emp["employee_id"], emp["name"], full_dates[di],
                        emp.get("shift", [])[di] if di < len(emp.get("shift", [])) else "",
                        emp.get("intime", [])[di] if di < len(emp.get("intime", [])) else "",
                        emp.get("outtime", [])[di] if di < len(emp.get("outtime", [])) else "",
                        emp.get("workhrs", [])[di] if di < len(emp.get("workhrs", [])) else "",
                    ])
    preview["OP"] = {"headers": op_h, "rows": op_rows, "total": count_op}

    # CL preview
    cl_h = ["EMPLOYEE_ID", "NAME", "DATE", "LEAVE_TYPE"]
    cl_rows = []
    count_cl = 0
    for emp in employees:
        leave_vals = emp.get("leave", [])
        for di in range(len(full_dates)):
            leave = leave_vals[di] if di < len(leave_vals) else None
            if leave and leave not in ("P", None, "WO"):
                count_cl += 1
                if len(cl_rows) < LIMIT:
                    cl_rows.append([emp["employee_id"], emp["name"], full_dates[di], leave])
    preview["CL"] = {"headers": cl_h, "rows": cl_rows, "total": count_cl}

    # DAT preview
    dat_h = ["EMPLOYEE_ID", "NAME", "DATE", "SHIFT", "LEAVE", "INTIME", "OUTTIME"]
    dat_rows = []
    total_dat = len(employees) * len([d for d in full_dates if d])
    for emp in employees[:3]:
        for di in range(min(7, len(full_dates))):
            if not full_dates[di]:
                continue
            dat_rows.append([
                emp["employee_id"], emp["name"], full_dates[di],
                emp.get("shift", [])[di] if di < len(emp.get("shift", [])) else "",
                emp.get("leave", [])[di] if di < len(emp.get("leave", [])) else "",
                emp.get("intime", [])[di] if di < len(emp.get("intime", [])) else "",
                emp.get("outtime", [])[di] if di < len(emp.get("outtime", [])) else "",
            ])
    preview["DAT"] = {"headers": dat_h, "rows": dat_rows, "total": total_dat}

    # A&B preview
    ab_h = ["EMPLOYEE_ID", "NAME"] + full_dates[:8] + ["..."]
    ab_rows = []
    for emp in employees[:LIMIT]:
        row = [emp["employee_id"], emp["name"]]
        leave_vals = emp.get("leave", [])
        shift_vals = emp.get("shift", [])
        for di in range(min(8, len(full_dates))):
            leave = leave_vals[di] if di < len(leave_vals) else None
            shift = shift_vals[di] if di < len(shift_vals) else None
            if leave and leave not in ("P", None):
                row.append(leave)
            elif shift:
                row.append(shift)
            else:
                row.append("")
        row.append("...")
        ab_rows.append(row)
    preview["A&B"] = {"headers": ab_h, "rows": ab_rows, "total": len(employees)}

    return preview


# ── Run ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=5000)
